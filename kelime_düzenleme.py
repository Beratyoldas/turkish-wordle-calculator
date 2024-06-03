import openpyxl



dosya=openpyxl.load_workbook("./kelime data.xlsx")
sayfa=dosya["Sayfa1"]
excel_satir_sayisi=sayfa.max_row



kelimeler = []
yenilenmis_kelimeler = []


for i in range(2,excel_satir_sayisi+1):
    kelime=sayfa.cell(i,1).value
    kelimeler.append(kelime)





for kelime in kelimeler:
    if isinstance(kelime, str):  # Eğer kelime bir string ise
        bosluk_index = kelime.find(" ")
        if bosluk_index != -1:
            yenilenmis_kelimeler.append(kelime[:bosluk_index])
        else:
            yenilenmis_kelimeler.append(kelime.strip())  # Boşlukları kaldır
    else:
        yenilenmis_kelimeler.append(kelime)




